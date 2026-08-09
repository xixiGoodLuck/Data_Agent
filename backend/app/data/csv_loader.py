from __future__ import annotations

import csv
import io
import json
import numbers
import re
import sqlite3
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.core.errors import AppError
from app.data.schema_reader import inspect_database
from app.models import Dataset

SQL_RESERVED = {
    "select",
    "from",
    "where",
    "group",
    "order",
    "limit",
    "table",
    "index",
    "join",
    "by",
    "as",
    "and",
    "or",
}
SUPPORTED_UPLOAD_EXTENSIONS = {".csv", ".xlsx"}
XLSX_UNCOMPRESSED_MIN_LIMIT = 10 * 1024 * 1024
XLSX_UNCOMPRESSED_RATIO = 10


@dataclass
class ParsedTable:
    headers: list[str]
    frame: pd.DataFrame
    source_type: str
    sheet_name: str | None = None
    preserve_cell_types: bool = False


def sanitize_column_names(headers: list[str]) -> tuple[list[str], list[dict[str, str]]]:
    used: dict[str, int] = {}
    sanitized: list[str] = []
    mapping: list[dict[str, str]] = []
    for index, original in enumerate(headers, start=1):
        base = re.sub(r"[^a-zA-Z0-9_]+", "_", original.strip()).strip("_").lower()
        if not base:
            base = f"column_{index}"
        if base[0].isdigit():
            base = f"column_{base}"
        if base in SQL_RESERVED or base.startswith("sqlite_"):
            base = f"column_{base}"
        count = used.get(base, 0) + 1
        used[base] = count
        name = base if count == 1 else f"{base}_{count}"
        while name in sanitized:
            count += 1
            used[base] = count
            name = f"{base}_{count}"
        sanitized.append(name)
        mapping.append({"original": original, "sanitized": name})
    return sanitized, mapping


def _looks_like_header(headers: list[str], rows: list[list[Any]]) -> bool:
    if not headers or all(not item.strip() for item in headers):
        return False
    if not rows:
        return True
    numeric_header = all(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", item.strip()) for item in headers)
    return not numeric_header


def _column_type(values: pd.Series) -> tuple[str, bool]:
    clean = values.astype(str).str.strip()
    nonempty = clean[clean != ""]
    if nonempty.empty:
        return "TEXT", False
    integer_pattern = nonempty.str.fullmatch(r"[-+]?(?:0|[1-9]\d*)")
    if bool(integer_pattern.all()):
        return "INTEGER", False
    real_pattern = nonempty.str.fullmatch(r"[-+]?(?:\d+\.\d+|\d+)(?:[eE][-+]?\d+)?")
    if bool(real_pattern.all()):
        return "REAL", False
    date_pattern = nonempty.str.fullmatch(
        r"\d{4}-(?:0[1-9]|1[0-2])-(?:0[1-9]|[12]\d|3[01])(?:[T ][0-2]\d:[0-5]\d(?::[0-5]\d)?)?"
    )
    return "TEXT", bool(date_pattern.all())


def _is_empty_cell(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _stringify_cell(value: Any) -> str | None:
    if _is_empty_cell(value):
        return None
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _excel_column_type(values: pd.Series) -> tuple[str, bool]:
    nonempty = [value for value in values.tolist() if not _is_empty_cell(value)]
    if not nonempty:
        return "TEXT", False
    if all(isinstance(value, bool) for value in nonempty):
        return "BOOLEAN", False
    if all(
        isinstance(value, numbers.Integral) and not isinstance(value, bool) for value in nonempty
    ):
        return "INTEGER", False
    if all(isinstance(value, numbers.Real) and not isinstance(value, bool) for value in nonempty):
        return "REAL", False
    if all(
        isinstance(value, (date, datetime)) and not isinstance(value, bool) for value in nonempty
    ):
        return "TEXT", True
    if all(isinstance(value, str) for value in nonempty):
        text = pd.Series(nonempty, dtype="string")
        date_pattern = text.str.fullmatch(
            r"\d{4}-(?:0[1-9]|1[0-2])-(?:0[1-9]|[12]\d|3[01])(?:[T ][0-2]\d:[0-5]\d(?::[0-5]\d)?)?"
        )
        return "TEXT", bool(date_pattern.all())
    return "TEXT", False


def _normalize_dataframe(
    parsed: ParsedTable, settings: Settings
) -> tuple[pd.DataFrame, list[dict[str, str]], dict[str, str], dict[str, bool]]:
    headers = parsed.headers
    rows = parsed.frame.values.tolist()
    if len(headers) > settings.max_upload_columns:
        raise AppError(
            "invalid_upload",
            f"The file has more than {settings.max_upload_columns:,} columns.",
        )
    if parsed.frame.empty or all(all(_is_empty_cell(cell) for cell in row) for row in rows):
        raise AppError("invalid_upload", "The file contains no data rows.")
    if len(parsed.frame) > settings.max_upload_rows:
        raise AppError(
            "invalid_upload",
            f"The file has more than {settings.max_upload_rows:,} rows.",
            status_code=413,
        )
    if not _looks_like_header(headers, rows):
        raise AppError("invalid_upload", "The file appears to be missing a header row.")

    columns, mapping = sanitize_column_names(headers)
    frame = parsed.frame.copy()
    frame.columns = columns
    date_like: dict[str, bool] = {}
    sql_types: dict[str, str] = {}
    for column in columns:
        if parsed.preserve_cell_types:
            sql_type, is_date = _excel_column_type(frame[column])
        else:
            sql_type, is_date = _column_type(frame[column])
        sql_types[column] = sql_type
        date_like[column] = is_date
        values = frame[column].map(lambda value: None if _is_empty_cell(value) else value)
        if sql_type == "BOOLEAN":
            frame[column] = values.astype("boolean")
        elif sql_type == "INTEGER":
            frame[column] = pd.to_numeric(values, errors="raise").astype("Int64")
        elif sql_type == "REAL":
            frame[column] = pd.to_numeric(values, errors="raise")
        else:
            frame[column] = values.map(_stringify_cell)
    return frame, mapping, sql_types, date_like


def parse_csv(content: bytes) -> ParsedTable:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise AppError("invalid_upload", "The CSV must use UTF-8 encoding.") from exc

    try:
        reader = csv.reader(io.StringIO(text, newline=""), strict=True)
        raw_rows = list(reader)
    except csv.Error as exc:
        raise AppError("invalid_upload", "The CSV structure is malformed.") from exc
    if not raw_rows:
        raise AppError("invalid_upload", "The uploaded file is empty.")
    headers = raw_rows[0]
    rows = raw_rows[1:]
    if any(len(row) != len(headers) for row in rows):
        raise AppError("invalid_upload", "Every CSV row must have the same number of columns.")
    return ParsedTable(headers=headers, frame=pd.DataFrame(rows), source_type="csv_upload")


def _check_xlsx_archive(content: bytes, settings: Settings) -> None:
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            expanded_size = sum(item.file_size for item in archive.infolist())
    except (OSError, zipfile.BadZipFile) as exc:
        raise AppError("invalid_upload", "The XLSX workbook is corrupt or invalid.") from exc
    expanded_limit = max(
        XLSX_UNCOMPRESSED_MIN_LIMIT,
        settings.max_upload_bytes * XLSX_UNCOMPRESSED_RATIO,
    )
    if expanded_size > expanded_limit:
        raise AppError(
            "invalid_upload",
            "The XLSX workbook expands beyond the safe processing limit.",
            status_code=413,
        )


def _excel_cell_value(cell) -> Any:
    value = cell.value
    if isinstance(value, datetime) and cell.is_date:
        number_format = cell.number_format.lower()
        if "h" not in number_format and "s" not in number_format:
            return value.date()
    return value


def _worksheet_table(worksheet, settings: Settings) -> ParsedTable | None:
    # Sparse cells avoid expanding a worksheet because a distant empty cell was only formatted.
    populated = {
        coordinate: _excel_cell_value(cell)
        for coordinate, cell in worksheet._cells.items()
        if not _is_empty_cell(_excel_cell_value(cell))
    }
    if not populated:
        return None
    row_numbers = sorted({row for row, _column in populated})
    column_numbers = sorted({_column for _row, _column in populated})
    if len(row_numbers) < 2:
        return None
    if len(column_numbers) > settings.max_upload_columns:
        raise AppError(
            "invalid_upload",
            f"The file has more than {settings.max_upload_columns:,} columns.",
        )
    data_row_numbers = row_numbers[1:]
    if len(data_row_numbers) > settings.max_upload_rows:
        raise AppError(
            "invalid_upload",
            f"The file has more than {settings.max_upload_rows:,} rows.",
            status_code=413,
        )
    header_row = row_numbers[0]
    headers = [
        ""
        if _is_empty_cell(_excel_cell_value(worksheet.cell(header_row, column)))
        else str(_excel_cell_value(worksheet.cell(header_row, column)))
        for column in column_numbers
    ]
    rows = [
        [_excel_cell_value(worksheet.cell(row, column)) for column in column_numbers]
        for row in data_row_numbers
    ]
    return ParsedTable(
        headers=headers,
        frame=pd.DataFrame(rows),
        source_type="excel_upload",
        sheet_name=worksheet.title,
        preserve_cell_types=True,
    )


def parse_excel(content: bytes, settings: Settings) -> ParsedTable:
    _check_xlsx_archive(content, settings)
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=False,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise AppError("invalid_upload", "The XLSX workbook is corrupt or invalid.") from exc
    try:
        try:
            for worksheet in workbook.worksheets:
                parsed = _worksheet_table(worksheet, settings)
                if parsed is not None:
                    return parsed
        except AppError:
            raise
        except Exception as exc:
            raise AppError("invalid_upload", "The XLSX workbook could not be parsed.") from exc
    finally:
        workbook.close()
    raise AppError("invalid_upload", "The XLSX workbook contains no data rows.")


def _safe_source_filename(filename: str, extension: str) -> str:
    basename = Path(filename.replace("\\", "/")).name
    cleaned = re.sub(r"[\x00-\x1f\x7f]+", "", basename).strip()[:240]
    return cleaned or f"upload{extension}"


def _persist_dataset(
    *,
    session: Session,
    settings: Settings,
    filename: str,
    extension: str,
    content: bytes,
    parsed: ParsedTable,
) -> Dataset:
    frame, mapping, sql_types, date_like = _normalize_dataframe(parsed, settings)
    settings.ensure_runtime_dirs()
    dataset_id = str(uuid4())
    db_path = (settings.datasets_dir / f"{dataset_id}.sqlite3").resolve()
    upload_path = (settings.uploads_dir / f"{dataset_id}{extension}").resolve()
    source_filename = _safe_source_filename(filename, extension)
    try:
        upload_path.write_bytes(content)
        connection = sqlite3.connect(db_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            frame.to_sql("data", connection, if_exists="fail", index=False)
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

        schema = inspect_database(db_path)
        for column in schema["data"]["columns"]:
            column["date_like"] = date_like[column["name"]]
            column["type"] = sql_types[column["name"]]
        display_stem = source_filename.rsplit(".", 1)[0]
        display_name = display_stem.strip()[:120]
        source_label = "Excel" if parsed.source_type == "excel_upload" else "CSV"
        sheet_text = f' worksheet "{parsed.sheet_name}"' if parsed.sheet_name else ""
        summary = (
            f"Uploaded {source_label}{sheet_text} with {len(frame)} rows "
            f"and {len(frame.columns)} columns."
        )
        description = json.dumps(
            {
                "kind": "upload",
                "summary": summary,
                "source_filename": source_filename,
                "sheet_name": parsed.sheet_name,
            },
            ensure_ascii=False,
        )
        dataset = Dataset(
            id=dataset_id,
            name=display_name or "Uploaded dataset",
            description=description,
            source_type=parsed.source_type,
            db_path=str(db_path),
            tables_json='["data"]',
            schema_json=json.dumps(schema),
            column_mapping_json=json.dumps(mapping),
            row_count=len(frame),
            is_builtin=False,
        )
        session.add(dataset)
        session.flush()
        return dataset
    except Exception:
        db_path.unlink(missing_ok=True)
        upload_path.unlink(missing_ok=True)
        raise


def ingest_csv(
    *,
    session: Session,
    settings: Settings,
    filename: str,
    content: bytes,
) -> Dataset:
    return ingest_tabular_file(
        session=session,
        settings=settings,
        filename=filename,
        content=content,
    )


def ingest_tabular_file(
    *,
    session: Session,
    settings: Settings,
    filename: str,
    content: bytes,
) -> Dataset:
    extension = Path(filename.replace("\\", "/")).suffix.lower()
    if extension not in SUPPORTED_UPLOAD_EXTENSIONS:
        raise AppError("invalid_upload", "Only .csv and .xlsx files are supported.")
    if not content:
        raise AppError("invalid_upload", "The uploaded file is empty.")
    if len(content) > settings.max_upload_bytes:
        raise AppError(
            "invalid_upload",
            f"The upload exceeds the {settings.max_upload_bytes:,} byte limit.",
            status_code=413,
        )
    parsed = parse_csv(content) if extension == ".csv" else parse_excel(content, settings)
    return _persist_dataset(
        session=session,
        settings=settings,
        filename=filename,
        extension=extension,
        content=content,
        parsed=parsed,
    )


def delete_uploaded_dataset(dataset: Dataset, settings: Settings) -> None:
    if dataset.is_builtin or dataset.source_type not in {"csv_upload", "excel_upload"}:
        raise AppError("invalid_upload", "Built-in datasets cannot be deleted.", status_code=409)
    db_path = Path(dataset.db_path).resolve()
    if db_path.parent != settings.datasets_dir.resolve():
        raise AppError("invalid_upload", "Dataset storage path is invalid.")
    db_path.unlink(missing_ok=True)
    extension = ".xlsx" if dataset.source_type == "excel_upload" else ".csv"
    (settings.uploads_dir / f"{dataset.id}{extension}").resolve().unlink(missing_ok=True)
