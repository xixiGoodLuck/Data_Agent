from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.core.config import Settings
from app.main import create_app


def xlsx_bytes(configure) -> bytes:
    workbook = Workbook()
    configure(workbook)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_xlsx_upload_creates_preview_schema_and_source_metadata(
    client: TestClient, test_settings: Settings
) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.title = "Data"
        sheet.append(["name", "region", "revenue", "code", "active"])
        sheet.append(["A", "East", 100, "00123", True])
        sheet.append(["B", "West", 200.5, "00456", False])
        sheet["Z10000"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    response = client.post(
        "/api/datasets/upload",
        files={
            "file": (
                "metrics.xlsx",
                xlsx_bytes(configure),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    body = response.json()
    assert body["source_type"] == "excel_upload"
    assert body["source_filename"] == "metrics.xlsx"
    assert body["sheet_name"] == "Data"
    assert body["row_count"] == 2
    assert body["preview"] == [
        {"name": "A", "region": "East", "revenue": 100.0, "code": "00123", "active": 1},
        {"name": "B", "region": "West", "revenue": 200.5, "code": "00456", "active": 0},
    ]
    assert body["schema"]["data"]["columns"][2]["type"] == "REAL"
    assert body["schema"]["data"]["columns"][3]["type"] == "TEXT"
    assert body["schema"]["data"]["columns"][4]["type"] == "BOOLEAN"
    summary = next(item for item in client.get("/api/datasets").json() if item["id"] == body["id"])
    assert summary["source_filename"] == "metrics.xlsx"
    assert summary["sheet_name"] == "Data"
    assert (test_settings.uploads_dir / f"{body['id']}.xlsx").exists()
    assert client.delete(f"/api/datasets/{body['id']}").status_code == 200
    assert not (test_settings.uploads_dir / f"{body['id']}.xlsx").exists()


def test_xlsx_upload_uses_first_nonempty_sheet_and_preserves_chinese_headers(
    client: TestClient,
) -> None:
    headers = ["\u5730\u533a", "\u9500\u552e\u989d", "\u65e5\u671f", "select", "\u9500\u552e\u989d"]

    def configure(workbook: Workbook) -> None:
        empty = workbook.active
        empty.title = "Empty"
        empty["A5000"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        sheet = workbook.create_sheet("\u660e\u7ec6")
        sheet.append(headers)
        sheet.append(["\u534e\u4e1c", 100, date(2026, 8, 9), "ok", 1])
        sheet.append(["\u534e\u5357", 200.5, datetime(2026, 8, 10, 12, 30), "ok", None])

    response = client.post(
        "/api/datasets/upload",
        files={"file": ("\u4e2d\u6587.xlsx", xlsx_bytes(configure), "application/octet-stream")},
    )

    assert response.status_code == 201
    body = response.json()
    assert body["sheet_name"] == "\u660e\u7ec6"
    assert [item["original"] for item in body["column_mapping"]] == headers
    assert [item["sanitized"] for item in body["column_mapping"]] == [
        "column_1",
        "column_2",
        "column_3",
        "column_select",
        "column_5",
    ]
    assert body["preview"][0]["column_3"] == "2026-08-09"
    assert body["preview"][1]["column_3"] == "2026-08-10T12:30:00"
    assert body["preview"][1]["column_5"] is None
    schema = body["schema"]["data"]["columns"]
    assert schema[1]["type"] == "REAL"
    assert schema[2]["type"] == "TEXT"
    assert schema[2]["date_like"] is True


@pytest.mark.parametrize(
    "filename", ["test.xls", "test.xlsm", "test.xltm", "test.xlsb", "test.txt"]
)
def test_upload_rejects_unsupported_spreadsheet_extensions(
    client: TestClient, filename: str
) -> None:
    response = client.post(
        "/api/datasets/upload",
        files={"file": (filename, b"not-supported", "application/octet-stream")},
    )

    assert response.status_code == 400
    assert response.json()["error"]["type"] == "invalid_upload"
    assert response.json()["error"]["message"] == "Only .csv and .xlsx files are supported."


@pytest.mark.parametrize(
    "filename,content",
    [
        ("empty.xlsx", b""),
        ("corrupt.xlsx", b"this is not an xlsx workbook"),
        ("empty-workbook.xlsx", xlsx_bytes(lambda _workbook: None)),
    ],
)
def test_xlsx_upload_rejects_empty_and_corrupt_workbooks(
    client: TestClient, filename: str, content: bytes
) -> None:
    response = client.post(
        "/api/datasets/upload",
        files={"file": (filename, content, "application/octet-stream")},
    )

    assert response.status_code in {400, 413}
    assert response.json()["error"]["type"] == "invalid_upload"


def test_xlsx_upload_rejects_missing_header(client: TestClient) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.append([1, 2])
        sheet.append([3, 4])

    response = client.post(
        "/api/datasets/upload",
        files={"file": ("no-header.xlsx", xlsx_bytes(configure), "application/octet-stream")},
    )

    assert response.status_code == 400
    assert response.json()["error"]["type"] == "invalid_upload"


def test_xlsx_upload_enforces_row_and_column_limits(tmp_path) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.append(["a", "b", "c"])
        sheet.append([1, 2, 3])
        sheet.append([4, 5, 6])

    content = xlsx_bytes(configure)
    row_settings = Settings(
        runtime_dir=tmp_path / "rows-runtime",
        max_upload_rows=1,
        log_level="WARNING",
    )
    column_settings = Settings(
        runtime_dir=tmp_path / "columns-runtime",
        max_upload_columns=2,
        log_level="WARNING",
    )
    with TestClient(create_app(row_settings)) as limited:
        response = limited.post(
            "/api/datasets/upload",
            files={"file": ("rows.xlsx", content, "application/octet-stream")},
        )
        assert response.status_code == 413
        assert response.json()["error"]["type"] == "invalid_upload"
    with TestClient(create_app(column_settings)) as limited:
        response = limited.post(
            "/api/datasets/upload",
            files={"file": ("columns.xlsx", content, "application/octet-stream")},
        )
        assert response.status_code == 400
        assert response.json()["error"]["type"] == "invalid_upload"


def test_xlsx_formula_without_cached_value_becomes_null(client: TestClient) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.append(["label", "amount"])
        sheet.append(["A", 10])
        sheet.append(["calculated", "=SUM(B2:B2)"])

    response = client.post(
        "/api/datasets/upload",
        files={"file": ("formula.xlsx", xlsx_bytes(configure), "application/octet-stream")},
    )

    assert response.status_code == 201
    assert response.json()["preview"][1] == {"label": "calculated", "amount": None}


def test_xlsx_mixed_type_column_safely_degrades_to_text(client: TestClient) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.append(["identifier"])
        sheet.append([123])
        sheet.append(["00123"])

    response = client.post(
        "/api/datasets/upload",
        files={"file": ("mixed.xlsx", xlsx_bytes(configure), "application/octet-stream")},
    )

    assert response.status_code == 201
    body = response.json()
    assert body["schema"]["data"]["columns"][0]["type"] == "TEXT"
    assert body["preview"] == [{"identifier": "123"}, {"identifier": "00123"}]


def test_xlsx_upload_rejects_unsafe_expanded_archive(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.append(["name"])
        sheet.append(["A"])

    monkeypatch.setattr("app.data.csv_loader.XLSX_UNCOMPRESSED_MIN_LIMIT", 100)
    monkeypatch.setattr("app.data.csv_loader.XLSX_UNCOMPRESSED_RATIO", 0)
    response = client.post(
        "/api/datasets/upload",
        files={"file": ("expanded.xlsx", xlsx_bytes(configure), "application/octet-stream")},
    )

    assert response.status_code == 413
    assert response.json()["error"]["type"] == "invalid_upload"


def test_uploaded_xlsx_uses_existing_agent_query_graph(client: TestClient) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.title = "Revenue"
        sheet.append(["product", "revenue"])
        sheet.append(["A", 100])
        sheet.append(["B", 300])

    uploaded = client.post(
        "/api/datasets/upload",
        files={"file": ("revenue.xlsx", xlsx_bytes(configure), "application/octet-stream")},
    ).json()
    response = client.post(
        "/api/query",
        json={"dataset_id": uploaded["id"], "question": "What is the total revenue?"},
    )

    assert response.status_code == 200
    assert response.json()["status"] == "success"
    assert response.json()["safe_sql"] is True
    assert response.json()["rows"][0]["total_revenue"] == 400
